import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def normalizar_cedula(val):
    """
    Convierte cualquier valor de celda en una cédula de 10 dígitos limpia.
    Ejemplo: 1712345678.0 -> '1712345678' | 12345678 -> '0012345678'
    """
    if pd.isna(val) or str(val).strip() == "":
        return None
    # Eliminar decimales de Excel (ej: .0) y caracteres no numéricos
    limpia = "".join(filter(str.isdigit, str(val).split('.')[0]))
    if not limpia:
        return None
    return limpia.zfill(10)[:10]

def unificar_excels_docentes(bytes_excel_titulos, bytes_excel_materias):
    """
    Cruza el Excel 1 (Títulos) con el Excel 2 (Materias).
    Expande a 1 fila por materia (1 a N) y prepara las columnas H, I y J para la IA.
    """
    print("--> [UNIFICACIÓN] Iniciando lectura y limpieza de archivos Excel...")
    
    # =========================================================================
    # 1. LECTURA Y LIMPIEZA EXCEL 1 (TÍTULOS - Desde Fila 3)
    # =========================================================================
    df_titulos = pd.read_excel(io.BytesIO(bytes_excel_titulos), header=None, skiprows=2)
    
    # Asignamos nombres temporales a las columnas clave del Excel 1
    # Col C = índice 2 | Col F = índice 5 | Col G = índice 6
    df_titulos['cedula_clean'] = df_titulos[2].apply(normalizar_cedula)
    
    # Eliminamos filas que no tengan cédula válida (filas vacías o subtítulos de tabla)
    df_titulos = df_titulos.dropna(subset=['cedula_clean']).copy()
    
    print(f"    [EXCEL 1] Se encontraron {len(df_titulos)} docentes únicos con títulos.")

    # =========================================================================
    # 2. LECTURA Y LIMPIEZA EXCEL 2 (MATERIAS / CARGA HORARIA)
    # =========================================================================
    df_materias = pd.read_excel(io.BytesIO(bytes_excel_materias), header=None)
    
    # Col C = índice 2 (Cédula) | Col L = índice 11 (Celda combinada L, M, N)
    df_materias['cedula_clean'] = df_materias[2].apply(normalizar_cedula)
    df_materias['materia_clean'] = df_materias[11].astype(str).str.strip()
    
    # Filtrar filas vacías, encabezados o materias nulas
    df_materias = df_materias.dropna(subset=['cedula_clean']).copy()
    df_materias = df_materias[df_materias['materia_clean'] != 'nan']
    df_materias = df_materias[df_materias['materia_clean'] != '']
    
    # Nos quedamos solo con Cédula y Materia del Excel 2
    df_materias_simple = df_materias[['cedula_clean', 'materia_clean']].copy()
    
    print(f"    [EXCEL 2] Se encontraron {len(df_materias_simple)} asignaciones de materias en total.")

    # =========================================================================
    # 3. CRUCE RELACIONAL (MERGE 1 a N)
    # =========================================================================
    df_unificado = pd.merge(df_titulos, df_materias_simple, on='cedula_clean', how='inner')
    
    print(f"--> [CRUCE EXITOSO] Matriz expandida a {len(df_unificado)} filas (Docente x Materia).")

    # =========================================================================
    # 4. CONSTRUCCIÓN DEL EXCEL DE SALIDA (ESTRUCTURA EXACTA)
    # =========================================================================
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Distributivo_Unificado"
    
    # Estilos profesionales para encabezados
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid") # Azul Marino Institucional
    align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Fila 1 y 2: Encabezados (Mantenemos estructura A-G y agregamos H, I, J)
    encabezados = [
        "Nro.", "Facultad / Carrera", "Cédula", "Nombres y Apellidos", 
        "Tipo Docente", "Título 3er Nivel (Grado)", "Título 4to Nivel (Posgrado)",
        "Materia Asignada (Col H)", "Observación IA (Col I)", "Análisis Técnico (Col J)"
    ]
    
    # Escribir fila de encabezados en la fila 2 para dejar estética limpia
    for col_idx, text in enumerate(encabezados, start=1):
        cell = ws_out.cell(row=2, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_thin
    
    ws_out.row_dimensions[2].height = 28

    # Inyectar los datos fila por fila desde la fila 3
    fila_actual = 3
    for idx, row in df_unificado.iterrows():
        # Columnas A hasta G (las primeras 7 columnas originales del Excel 1)
        for col_orig in range(7):
            val = row[col_orig] if col_orig in row else ""
            cell = ws_out.cell(row=fila_actual, column=col_orig+1, value=val)
            cell.border = border_thin
            if col_orig == 2: # Alinear cédula al centro
                cell.alignment = Alignment(horizontal="center")
                
        # Columna H: Materia (proviene de la columna L del Excel 2)
        cell_h = ws_out.cell(row=fila_actual, column=8, value=row['materia_clean'])
        cell_h.border = border_thin
        
        # Columna I: Observación IA (Vacía por ahora, lista para Gemini)
        cell_i = ws_out.cell(row=fila_actual, column=9, value="")
        cell_i.border = border_thin
        cell_i.alignment = Alignment(horizontal="center")
        
        # Columna J: Análisis técnico (Vacío por ahora, listo para Gemini)
        cell_j = ws_out.cell(row=fila_actual, column=10, value="")
        cell_j.border = border_thin
        
        fila_actual += 1

    # Ajustar ancho de columnas automáticamente para legibilidad
    anchos = {"A": 6, "B": 20, "C": 13, "D": 32, "E": 15, "F": 38, "G": 38, "H": 35, "I": 18, "J": 50}
    for col_letra, ancho in anchos.items():
        ws_out.column_dimensions[col_letra].width = ancho

    # Exportar a bytes
    output_buffer = io.BytesIO()
    wb_out.save(output_buffer)
    print("--> [TERMINADO] Archivo unificado generado correctamente.")
    return output_buffer.getvalue()