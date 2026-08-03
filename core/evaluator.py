import io
import json
import re
import time
import unicodedata
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font

def limpiar_json_md(texto_bruto):
    """Elimina bloques Markdown si el LLM los incluye."""
    texto = re.sub(r'^```json\s*', '', texto_bruto, flags=re.MULTILINE | re.IGNORECASE)
    texto = re.sub(r'^```\s*', '', texto, flags=re.MULTILINE)
    return texto.strip()

def normalizar_materia(texto):
    """
    Convierte a mayúsculas, quita tildes y elimina cualquier carácter
    que no sea una letra o número (espacios, puntos, comas, guiones).
    Ideal para hacer matches exactos aunque la IA 'limpie' el texto.
    """
    if not texto:
        return ""
    t = str(texto).upper()
    # Quitar tildes
    t = ''.join(c for c in unicodedata.normalize('NFD', t) if unicodedata.category(c) != 'Mn')
    # Quitar todo lo que no sea alfanumérico
    t = re.sub(r'[^A-Z0-9]', '', t)
    return t

def auditar_docente_ia(client, titulos_3er, titulos_4to, materias, modelo_nombre):
    prompt = f"""
    Eres un par evaluador académico senior y auditor de la SENESCYT / CES en Ecuador.
    Tu tarea es auditar la idoneidad y compatibilidad curricular de un docente universitario para impartir las materias asignadas según la normativa de Educación Superior (LOES).

    PERFIL DEL DOCENTE:
    - Título(s) de 3er Nivel (Grado): "{titulos_3er}"
    - Título(s) de 4to Nivel (Posgrado / Maestría / Doctorado): "{titulos_4to}"

    MATERIAS ASIGNADAS A EVALUAR (Listado):
    {json.dumps(materias, ensure_ascii=False)}

    REGLAS ESTRICTAS DE EVALUACIÓN:
    1. Para enseñar en tercer nivel, el docente debe tener preferentemente un título de posgrado afín a la cátedra, o en su defecto, un título de grado directamente afín con amplia especialización.
    2. Si los títulos son totalmente ajenos al área de la materia (Ejemplo: Máster en Finanzas dictando Anatomía Humana), debes marcar "No cumple".
    3. Si existe afinidad directa y lógica, marca "Cumple".
    4. Si es una materia multidisciplinaria o de frontera donde la afinidad es debatible y requiere revisión del comité académico, marca "Verificar".

    FORMATO DE RESPUESTA OBLIGATORIO (JSON ESTRICTO):
    Debes responder ÚNICAMENTE con un objeto JSON que contenga la llave "evaluaciones" con el arreglo de materias.
    Estructura exacta:
    {{
      "evaluaciones": [
        {{"materia": "MATEMATICAS I", "observacion": "Cumple", "analisis": "El título de Ingeniero Civil acredita solvencia analítica."}},
        {{"materia": "ANATOMIA", "observacion": "No cumple", "analisis": "Su formación en Administración no guarda relación con salud."}}
      ]
    }}
    """
    
    try:
        chat_completion = client.chat.completions.create(
            messages=[
                {
                    "role": "system",
                    "content": "Eres un auditor académico formal. Respondes SIEMPRE en formato JSON válido."
                },
                {
                    "role": "user",
                    "content": prompt,
                }
            ],
            model=modelo_nombre,
            temperature=0.1,
            response_format={"type": "json_object"}
        )
        
        texto_limpio = chat_completion.choices[0].message.content
        texto_limpio = limpiar_json_md(texto_limpio)
        datos_json = json.loads(texto_limpio)
        
        lista_evaluaciones = datos_json.get("evaluaciones", [])
        if isinstance(datos_json, list):
            lista_evaluaciones = datos_json
            
        dict_evaluaciones = {}
        for item in lista_evaluaciones:
            # 👉 AQUÍ NORMALIZAMOS LA LLAVE DEL DICCIONARIO
            mat_original = str(item.get("materia", ""))
            key_mat = normalizar_materia(mat_original)
            
            dict_evaluaciones[key_mat] = {
                "observacion": item.get("observacion", "Verificar"),
                "analisis": item.get("analisis", "No se pudo generar justificación detallada.")
            }
        return dict_evaluaciones
        
    except Exception as e:
        print(f"[!] Error rápido en llamada a la API: {e}")
        # En caso de error, devolvemos un dict con las llaves normalizadas también
        return {normalizar_materia(mat): {"observacion": "Verificar", "analisis": f"Error en evaluación de IA: {str(e)}"} for mat in materias}

def evaluar_matriz_completa_ia(archivo_bytes, api_key, progress_callback, modelo_nombre="gpt-4o-mini"):
    
    if "gpt" in modelo_nombre.lower():
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        print(f"--> [CONEXIÓN API] Conectando a OpenAI (Modelo: {modelo_nombre})")
    else:
        from groq import Groq
        client = Groq(api_key=api_key)
        print(f"--> [CONEXIÓN API] Conectando a Groq (Modelo: {modelo_nombre})")
    
    wb = openpyxl.load_workbook(io.BytesIO(archivo_bytes))
    ws = wb.active
    
    fila_inicio = 3
    fila_fin = ws.max_row
    
    docentes_agrupados = {}
    print(f"--> [IA-AUDITORÍA] Agrupando filas por docente...")
    for num_fila in range(fila_inicio, fila_fin + 1):
        cedula = str(ws.cell(row=num_fila, column=3).value or "").strip()
        if not cedula:
            continue
            
        t_3er = str(ws.cell(row=num_fila, column=6).value or "No registra").strip()
        t_4to = str(ws.cell(row=num_fila, column=7).value or "No registra").strip()
        materia = str(ws.cell(row=num_fila, column=8).value or "").strip()
        
        if cedula not in docentes_agrupados:
            docentes_agrupados[cedula] = {
                'filas': [],
                '3er': t_3er,
                '4to': t_4to,
                'materias': []
            }
        
        docentes_agrupados[cedula]['filas'].append(num_fila)
        if materia and materia not in docentes_agrupados[cedula]['materias']:
            docentes_agrupados[cedula]['materias'].append(materia)
            
    total_docentes = len(docentes_agrupados)
    
    fill_cumple = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    fill_no_cumple = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    fill_verificar = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    font_bold = Font(name="Calibri", size=10, bold=True)
    
    docente_actual = 0
    for cedula, datos in docentes_agrupados.items():
        docente_actual += 1
        progress_callback(docente_actual / total_docentes, f"[{docente_actual}/{total_docentes}] Auditando Cédula {cedula}...")
        
        dict_respuestas = auditar_docente_ia(
            client=client,
            titulos_3er=datos['3er'],
            titulos_4to=datos['4to'],
            materias=datos['materias'],
            modelo_nombre=modelo_nombre 
        )
        
        for fila_idx in datos['filas']:
            mat_fila_original = str(ws.cell(row=fila_idx, column=8).value or "")
            
            # 👉 AQUÍ NORMALIZAMOS LO QUE VIENE DE EXCEL PARA QUE EL MATCH SEA PERFECTO
            key_mat_fila = normalizar_materia(mat_fila_original)
            
            eval_mat = dict_respuestas.get(key_mat_fila, {
                "observacion": "Verificar", 
                "analisis": "Materia no evaluada por el modelo."
            })
            
            obs_val = eval_mat["observacion"]
            analisis_val = eval_mat["analisis"]
            
            celda_i = ws.cell(row=fila_idx, column=9, value=obs_val)
            celda_i.font = font_bold
            celda_i.alignment = Alignment(horizontal="center", vertical="center")
            
            if "NO CUMPLE" in obs_val.upper():
                celda_i.fill = fill_no_cumple
            elif "CUMPLE" in obs_val.upper():
                celda_i.fill = fill_cumple
            else:
                celda_i.fill = fill_verificar
                
            celda_j = ws.cell(row=fila_idx, column=10, value=analisis_val)
            celda_j.alignment = Alignment(vertical="center", wrap_text=True)
            
        time.sleep(4.0)
        
    ws.row_dimensions[2].height = 28
    
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    return output_buffer.getvalue()