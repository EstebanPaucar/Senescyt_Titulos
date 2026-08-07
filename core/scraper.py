import time
import io
import os
import json
import openpyxl
import cv2
import pytesseract
from PIL import Image
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
import concurrent.futures
from collections import Counter
import sys
import asyncio

# Parche híbrido local/nube
if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
else:
    pytesseract.pytesseract.tesseract_cmd = '/usr/bin/tesseract'

ARCHIVO_CACHE = "historial_auditorias.json"

def cargar_cache():
    if os.path.exists(ARCHIVO_CACHE):
        try:
            with open(ARCHIVO_CACHE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def guardar_cache(data_cache):
    try:
        with open(ARCHIVO_CACHE, "w", encoding="utf-8") as f:
            json.dump(data_cache, f, ensure_ascii=False, indent=4)
    except Exception:
        pass

def resolver_captcha(ruta_imagen):
    try:
        img = cv2.imread(ruta_imagen, cv2.IMREAD_GRAYSCALE)
        img_agrandada = cv2.resize(img, None, fx=3.0, fy=3.0, interpolation=cv2.INTER_CUBIC)
        img_limpia = cv2.fastNlMeansDenoising(img_agrandada, None, h=15, templateWindowSize=7, searchWindowSize=21)
        img_invertida = cv2.bitwise_not(img_limpia)
        _, img_bin = cv2.threshold(img_invertida, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
        img_morf = cv2.morphologyEx(img_bin, cv2.MORPH_CLOSE, kernel)
        
        ruta_procesada = ruta_imagen.replace(".png", "_proc.png")
        cv2.imwrite(ruta_procesada, img_morf)
        
        config = r'--oem 3 --psm 7 -c tessedit_char_whitelist=abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'
        texto = pytesseract.image_to_string(Image.open(ruta_procesada), config=config)
        return "".join(filter(str.isalnum, texto.strip()))
    except Exception as e:
        print(f"Error OCR: {e}")
        return ""

def consultar_senescyt_web(cedula, id_hilo):
    print(f"[Hilo-{id_hilo}] Iniciando extracción para {cedula}...")
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
        page = context.new_page()
        
        muestras_obtenidas = []
        intentos = 0
        
        while len(muestras_obtenidas) < 3 and intentos < 6:
            intentos += 1
            try:
                page.goto("https://www.senescyt.gob.ec/web/guest/consultas", timeout=60000)
                page.wait_for_load_state('networkidle')
                
                try:
                    if page.locator('a.ui-dialog-titlebar-close').first.is_visible(timeout=1500):
                        page.locator('a.ui-dialog-titlebar-close').first.click()
                except:
                    pass
                
                selector_cedula = 'input[id="formPrincipal:identificacion"]'
                page.locator(selector_cedula).fill(cedula)
                
                ruta_cap = f"captcha_{id_hilo}.png"
                page.locator('img[id="formPrincipal:capimg"]').screenshot(path=ruta_cap)
                texto_captcha = resolver_captcha(ruta_cap)
                
                if not texto_captcha or len(texto_captcha) < 4 or len(texto_captcha) > 6:
                    time.sleep(1)
                    continue
                    
                page.locator('input[id="formPrincipal:captchaSellerInput"]').fill(texto_captcha)
                page.click('button[id="formPrincipal:boton-buscar"]')
                
                tiempo_inicio = time.time()
                tabla_lista = False
                
                while time.time() - tiempo_inicio < 9.5:
                    html_lower = page.content().lower()
                    if page.locator('tbody.ui-datatable-data').count() > 0:
                        tabla_lista = True
                        break
                    if any(x in html_lower for x in ["ningún registro", "no se encontr", "sin registro"]):
                        tabla_lista = True
                        break
                    time.sleep(0.5)
                
                if not tabla_lista:
                    continue
                    
                soup = BeautifulSoup(page.content(), 'html.parser')
                t3, t4 = [], []
                paneles = soup.find_all('h4', class_='panel-title')
                
                for enc in paneles:
                    txt_enc = enc.get_text(strip=True).lower()
                    tbody = enc.find_next('tbody', class_='ui-datatable-data')
                    if not tbody: continue
                    
                    filas = tbody.find_all('tr')
                    for f in filas:
                        tds = f.find_all('td')
                        if tds:
                            span = tds[0].find('span', class_='ui-column-title')
                            if span: span.decompose()
                            titulo = tds[0].get_text(separator=' ', strip=True)
                            
                            if "cuarto nivel" in txt_enc or "posgrado" in txt_enc:
                                if titulo not in t4: t4.append(titulo)
                            elif "tercer nivel" in txt_enc or "grado" in txt_enc:
                                if titulo not in t3: t3.append(titulo)
                
                res_3er = " | ".join(t3) if t3 else "No registra"
                res_4to = " | ".join(t4) if t4 else "No registra"
                muestras_obtenidas.append((res_3er, res_4to))
                
                if Counter(muestras_obtenidas).most_common(1)[0][1] >= 2:
                    break
                    
            except Exception as e:
                print(f"[Hilo-{id_hilo}] Error de red: {str(e)[:50]}")
                
        browser.close()
        
        if muestras_obtenidas:
            ganador = Counter(muestras_obtenidas).most_common(1)[0][0]
            print(f"[Hilo-{id_hilo}] ✅ Éxito {cedula}: T3({len(ganador[0][:15])}...)")
            return ganador[0], ganador[1]
        
        print(f"[Hilo-{id_hilo}] ❌ Fallo total {cedula}")
        return "No registra", "No registra"

def procesar_excel_concurrente(archivo_bytes):
    print("🚀 INICIANDO ORQUESTADOR MULTITHREADING (5 HILOS)")
    wb = openpyxl.load_workbook(io.BytesIO(archivo_bytes))
    ws = wb.active
    
    cache = cargar_cache()
    pendientes = []
    
    # Evaluar qué cédulas necesitan ser consultadas
    for fila in range(3, ws.max_row + 1):
        ced_cruda = ws.cell(row=fila, column=3).value
        if not ced_cruda: continue
        
        ced = ''.join(filter(str.isdigit, str(ced_cruda).split('.')[0])).zfill(10)[:10]
        
        if ced in cache and cache[ced].get("tercer_nivel") != "No registra":
            ws.cell(row=fila, column=6).value = cache[ced]["tercer_nivel"]
            ws.cell(row=fila, column=7).value = cache[ced]["cuarto_nivel"]
        else:
            pendientes.append((fila, ced))
            
    print(f"📊 Recuperados de caché: {ws.max_row - 2 - len(pendientes)} | A procesar en web: {len(pendientes)}")

    # Desplegar el Pool de Hilos
    if pendientes:
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            futuros = {executor.submit(consultar_senescyt_web, ced, i): (fila, ced) for i, (fila, ced) in enumerate(pendientes)}
            
            for futuro in concurrent.futures.as_completed(futuros):
                try:
                    fila, ced = futuros[futuro]
                    t3, t4 = futuro.result() # Si hay un error crítico en el hilo, saltará aquí
                    
                    # Escribir en la memoria del Excel
                    ws.cell(row=fila, column=6).value = t3
                    ws.cell(row=fila, column=7).value = t4
                    
                    # 💾 CHECKPOINT: Actualizar y guardar la caché físicamente
                    cache[ced] = {"tercer_nivel": t3, "cuarto_nivel": t4}
                    guardar_cache(cache)
                    
                except Exception as e:
                    print(f"⚠️ Error crítico procesando la fila {fila}: {e}")
        
        guardar_cache(cache)

    salida = io.BytesIO()
    wb.save(salida)
    print("✅ PROCESO COMPLETADO EXITOSAMENTE.")
    return salida.getvalue()